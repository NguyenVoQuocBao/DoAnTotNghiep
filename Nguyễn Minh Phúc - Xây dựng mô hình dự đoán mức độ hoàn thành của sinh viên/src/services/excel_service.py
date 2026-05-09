"""Enhanced Excel export service (simplified version)"""

import sys
import pandas as pd
from pathlib import Path
from typing import Optional, Tuple, Dict, List
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# Add parent directory to path
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from config import EXCEL_FILE, EXCEL_FALLBACK

class ExcelService:
    """Enhanced service for Excel operations (without PDF for now)"""
    
    _cache = None
    _mssv_mapping = None
    
    @classmethod
    def load_excel_data(cls) -> Tuple[Optional[pd.DataFrame], Dict]:
        """Load Excel data with caching"""
        if cls._cache is not None:
            return cls._cache, cls._mssv_mapping
        
        excel_file = EXCEL_FILE if EXCEL_FILE.exists() else EXCEL_FALLBACK
        
        if not excel_file.exists():
            return None, {}
        
        try:
            df = pd.read_excel(excel_file)
            cls._cache = df
            cls._mssv_mapping = cls._build_mssv_mapping(df)
            return df, cls._mssv_mapping
        except Exception as e:
            print(f"Error loading Excel: {e}")
            return None, {}
    
    @staticmethod
    def _build_mssv_mapping(df: pd.DataFrame) -> Dict:
        """Build MSSV to index mapping"""
        mapping = {}
        mssv_columns = [
            col for col in df.columns 
            if 'StudentID' in str(col) or 'MSSV' in str(col).upper() 
            or 'student_id' in str(col).lower()
        ]
        
        if mssv_columns:
            mssv_col = mssv_columns[0]
            for idx, row in df.iterrows():
                mssv = str(row.get(mssv_col, '')).strip()
                if mssv:
                    mapping[mssv.upper()] = idx
                    mapping[mssv] = idx
        
        return mapping
    
    @staticmethod
    def format_excel_file(file_path: Path):
        """Apply enhanced formatting to Excel file"""
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Header formatting
        header_fill = PatternFill(
            start_color="366092", 
            end_color="366092", 
            fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        # Border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center", 
                vertical="center", 
                wrap_text=True
            )
            cell.border = thin_border
        
        # Data formatting
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(file_path)
    
    @staticmethod
    def export_advanced_analysis(data: List[Dict], analytics_results: Dict) -> str:
        """Export comprehensive advanced analytics to Excel"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"advanced_analytics_report_{timestamp}.xlsx"
        filepath = Path(filename)
        
        # Create DataFrame
        df = pd.DataFrame(data)
        
        # Write to Excel with multiple sheets
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Main data sheet
            df.to_excel(writer, sheet_name='Student Analysis', index=False)
            
            # Summary statistics sheet
            summary_data = []
            if analytics_results.get('summary'):
                summary = analytics_results['summary']
                summary_data = [
                    ['Metric', 'Value'],
                    ['Total Students', summary.get('total_students', 0)],
                    ['High Performers (%)', f"{summary.get('high_performers_percentage', 0):.1f}%"],
                    ['At-Risk Students (%)', f"{summary.get('anomaly_rate', 0):.1f}%"],
                    ['Trend Direction', summary.get('trend_direction', 'Unknown').title()]
                ]
            
                summary_df = pd.DataFrame(summary_data[1:], columns=summary_data[0])
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Clustering results sheet
            if analytics_results.get('clustering'):
                cluster_data = []
                for cluster_id, cluster_info in analytics_results['clustering']['cluster_analysis'].items():
                    cluster_data.append({
                        'Cluster ID': cluster_id,
                        'Cluster Name': cluster_info['name'],
                        'Student Count': cluster_info['size'],
                        'Percentage': f"{cluster_info['percentage']:.1f}%",
                        'Avg Test Scores': f"{cluster_info['characteristics']['avg_test_scores']:.2f}",
                        'Avg Assignment Completion': f"{cluster_info['characteristics']['avg_assignment_completion']:.2f}",
                        'Avg Access Frequency': f"{cluster_info['characteristics']['avg_access_frequency']:.2f}"
                    })
                
                if cluster_data:
                    cluster_df = pd.DataFrame(cluster_data)
                    cluster_df.to_excel(writer, sheet_name='Clustering Analysis', index=False)
            
            # Anomalies sheet
            if analytics_results.get('anomalies') and analytics_results['anomalies'].get('anomalies'):
                anomaly_data = []
                for anomaly in analytics_results['anomalies']['anomalies']:
                    anomaly_data.append({
                        'Student ID': anomaly['student_id'],
                        'Anomaly Score': f"{anomaly['anomaly_score']:.3f}",
                        'Test Scores': f"{anomaly['features']['test_scores']:.2f}",
                        'Assignment Completion': f"{anomaly['features']['assignment_completion']:.2f}",
                        'Access Frequency': f"{anomaly['features']['access_frequency']:.2f}",
                        'Study Time': f"{anomaly['features']['study_time']:.2f}",
                        'Submission Timing': f"{anomaly['features']['submission_timing']:.2f}"
                    })
                
                if anomaly_data:
                    anomaly_df = pd.DataFrame(anomaly_data)
                    anomaly_df.to_excel(writer, sheet_name='At-Risk Students', index=False)
        
        # Apply formatting
        ExcelService.format_excel_file(filepath)
        
        return filename
    
    @staticmethod
    def export_to_pdf(data: List[Dict], analytics_results: Dict) -> str:
        """Export to PDF - placeholder for future implementation"""
        # For now, just create a simple text report
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"analytics_report_{timestamp}.txt"
        
        with open(filename, 'w', encoding='utf-8') as f:
            f.write("LEARNING ANALYTICS REPORT\n")
            f.write("=" * 50 + "\n\n")
            
            if analytics_results.get('summary'):
                summary = analytics_results['summary']
                f.write("SUMMARY\n")
                f.write("-" * 20 + "\n")
                f.write(f"Total Students: {summary.get('total_students', 0)}\n")
                f.write(f"High Performers: {summary.get('high_performers_percentage', 0):.1f}%\n")
                f.write(f"At-Risk Students: {summary.get('anomaly_rate', 0):.1f}%\n")
                f.write(f"Trend Direction: {summary.get('trend_direction', 'Unknown')}\n\n")
            
            f.write("STUDENT DATA\n")
            f.write("-" * 20 + "\n")
            for i, student in enumerate(data[:10], 1):  # First 10 students
                f.write(f"{i}. {student.get('Mã sinh viên', 'N/A')} - {student.get('Tên sinh viên', 'N/A')}\n")
                f.write(f"   Completion: {student.get('Mức độ hoàn thành', 'N/A')}\n")
                f.write(f"   Behavior: {student.get('Nhóm hành vi', 'N/A')}\n\n")
        
        return filename
    
    @classmethod
    def clear_cache(cls):
        """Clear Excel data cache"""
        cls._cache = None
        cls._mssv_mapping = None
    
    @staticmethod
    def load_external_excel(file_path: Path) -> Optional[pd.DataFrame]:
        """Load external Excel or CSV file"""
        try:
            if str(file_path).endswith('.csv'):
                df = pd.read_csv(file_path, encoding='utf-8-sig')
            else:
                df = pd.read_excel(file_path)
            return df
        except Exception as e:
            print(f"Error loading external file: {e}")
            return None
    
    @staticmethod
    def sort_excel_data(df: pd.DataFrame, sort_column: str, ascending: bool = True) -> pd.DataFrame:
        """Sort Excel data by specified column"""
        try:
            if sort_column in df.columns:
                return df.sort_values(by=sort_column, ascending=ascending)
            return df
        except Exception as e:
            print(f"Error sorting Excel data: {e}")
            return df
    
    @staticmethod
    def export_sorted_excel(df: pd.DataFrame, output_filename: str) -> Path:
        """Export sorted DataFrame to Excel with formatting"""
        output_path = Path(output_filename)
        df.to_excel(output_path, index=False, engine='openpyxl')
        ExcelService.format_excel_file(output_path)
        return output_path
    
    @staticmethod
    def filter_excel_data(df: pd.DataFrame, column: str, filter_value: str, filter_type: str = 'contains') -> pd.DataFrame:
        """Filter Excel data by column value
        
        Args:
            df: DataFrame to filter
            column: Column name to filter on
            filter_value: Value to filter by
            filter_type: Type of filter ('contains', 'equals', 'greater', 'less', 'starts_with', 'ends_with')
        """
        try:
            if column not in df.columns:
                return df
            
            if filter_type == 'contains':
                return df[df[column].astype(str).str.contains(filter_value, case=False, na=False)]
            elif filter_type == 'equals':
                return df[df[column].astype(str).str.lower() == filter_value.lower()]
            elif filter_type == 'starts_with':
                return df[df[column].astype(str).str.startswith(filter_value, na=False)]
            elif filter_type == 'ends_with':
                return df[df[column].astype(str).str.endswith(filter_value, na=False)]
            elif filter_type == 'greater':
                return df[pd.to_numeric(df[column], errors='coerce') > float(filter_value)]
            elif filter_type == 'less':
                return df[pd.to_numeric(df[column], errors='coerce') < float(filter_value)]
            else:
                return df
        except Exception as e:
            print(f"Error filtering data: {e}")
            return df
    
    @staticmethod
    def get_column_statistics(df: pd.DataFrame, column: str) -> Dict:
        """Get statistics for a numeric column"""
        try:
            if column not in df.columns:
                return {}
            
            # Try to convert to numeric
            numeric_data = pd.to_numeric(df[column], errors='coerce').dropna()
            
            if len(numeric_data) == 0:
                return {'type': 'non-numeric', 'unique_values': df[column].nunique()}
            
            return {
                'type': 'numeric',
                'count': len(numeric_data),
                'mean': round(numeric_data.mean(), 2),
                'median': round(numeric_data.median(), 2),
                'min': round(numeric_data.min(), 2),
                'max': round(numeric_data.max(), 2),
                'std': round(numeric_data.std(), 2)
            }
        except Exception as e:
            print(f"Error calculating statistics: {e}")
            return {}
    
    @staticmethod
    def export_to_csv(df: pd.DataFrame, output_filename: str) -> Path:
        """Export DataFrame to CSV"""
        output_path = Path(output_filename)
        df.to_csv(output_path, index=False, encoding='utf-8-sig')
        return output_path
    
    @staticmethod
    def export_to_json(df: pd.DataFrame, output_filename: str) -> Path:
        """Export DataFrame to JSON"""
        output_path = Path(output_filename)
        df.to_json(output_path, orient='records', force_ascii=False, indent=2)
        return output_path
    
    @staticmethod
    def import_from_csv(file_path: Path) -> Optional[pd.DataFrame]:
        """Import data from CSV file"""
        try:
            df = pd.read_csv(file_path, encoding='utf-8-sig')
            return df
        except Exception as e:
            print(f"Error importing CSV: {e}")
            return None
    
    @staticmethod
    def import_from_json(file_path: Path) -> Optional[pd.DataFrame]:
        """Import data from JSON file"""
        try:
            df = pd.read_json(file_path, orient='records')
            return df
        except Exception as e:
            print(f"Error importing JSON: {e}")
            return None
